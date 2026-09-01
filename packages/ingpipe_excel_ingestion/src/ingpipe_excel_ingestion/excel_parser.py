"""Generic Excel parser: explicit row/column bounds with sane defaults.

Reads Excel workbooks with openpyxl. The caller may supply 1-based row numbers
(``header_row``, ``data_start_row``, ``data_end_row``) and an Excel-letter column
span (``start_col``/``end_col``); ALL are optional and default sensibly:

  - ``header_row``     -> 1
  - ``data_start_row`` -> ``header_row + 1``
  - ``data_end_row``   -> the last row with content in the column span
                          (best-effort; includes trailing footer/marker rows, so
                          sheets with a footer below the data should set it)
  - ``start_col``      -> "A"
  - ``end_col``        -> the column of the first blank header from ``start_col``

There is NO header-anchor / stop-marker / skip-marker detection. All cell values
are coerced to stripped strings (empty -> ``None``). Merged ranges are resolved
(the anchor cell's value is carried onto every spanned cell, so a vertically
merged category column keeps its value on every row). Rows inside the data
range whose every value is empty are SKIPPED (logged with a count) rather than
ingested as all-``None`` records. Each kept data row carries a synthetic
1-based, contiguous ``row_number``.

A sheet is REJECTED (ValueError) if any header in the resolved column span is
missing (blank) or if two headers collide after cleaning (identical once
normalized to a ``col_*`` name), so a malformed header row is never silently
misaligned or collapsed.
"""

import posixpath
import re
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Sequence
from pathlib import Path

from ingpipe_lib.logconfig import get_logger
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

from ingpipe_excel_ingestion._utils import normalize_column_name

logger = get_logger(__name__)

# OOXML namespaces used to locate a sheet's XML part and its mergeCells
# entries. Merged-range metadata is read straight from the archive because
# openpyxl's read-only worksheets do not expose merged_cells, and the
# read-only pass is kept for its memory profile (a full load_workbook would
# materialize whole large workbooks).
_SPREADSHEET_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_PKG_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_DOC_REL_ID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"

# Key under which the synthetic positional ordinal is attached to each row dict.
# Kept distinct from any header so it never appears in row_text or the col_* map.
ROW_NUMBER_KEY = "row_number"

# An Excel column label is one or more ASCII letters (A, B, ..., Z, AA, ...).
_COLUMN_LETTER_RE = re.compile(r"[A-Za-z]+")

# Defaults for the optional bounds (data_end_row / end_col are resolved from the
# sheet content once it is loaded — see parse_sheet).
_DEFAULT_HEADER_ROW = 1
_DEFAULT_START_COL = "A"


def validate_column_letter(col: str, label: str) -> str:
    """Validate that a value is an Excel column label (e.g. ``"A"``, ``"AA"``).

    Args:
        col: The candidate column label.
        label: Field name for the error message (e.g. ``"start_col"``).

    Returns:
        ``col`` unchanged when valid.

    Raises:
        ValueError: If ``col`` is not one or more ASCII letters.
    """
    if not isinstance(col, str) or not _COLUMN_LETTER_RE.fullmatch(col):
        raise ValueError(
            f"{label} ({col!r}) must be an Excel column letter (e.g. 'A', 'B', 'AA')"
        )
    return col


def _column_index(col: str, label: str) -> int:
    """Return the 0-based column index for a validated Excel column label."""
    validate_column_letter(col, label)
    return column_index_from_string(col.upper()) - 1


def list_sheets(filepath: str | Path) -> list[str]:
    """Return all sheet names from an Excel workbook in workbook order.

    There is no exclusion logic: sheets are selected for processing by being
    listed in the config, so an unlisted sheet is simply never parsed.

    Args:
        filepath: Path to the Excel workbook.

    Returns:
        List of sheet names in workbook order.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the workbook cannot be opened (corrupt / invalid file).
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    logger.info(f"Listing sheets in {filepath.name}")

    wb = None
    try:
        wb = load_workbook(filepath, read_only=True)
    except (OSError, ValueError, zipfile.BadZipFile, InvalidFileException) as e:
        # A corrupt/unreadable workbook can surface as BadZipFile or
        # InvalidFileException (neither a ValueError); wrap so callers have one
        # exception type to record as a failure rather than crashing the run.
        logger.error(f"Failed to open workbook: {filepath} - {e}")
        raise ValueError(f"Cannot open workbook {filepath.name}: {e}") from e
    else:
        sheets = list(wb.sheetnames)
        logger.debug(f"Found {len(sheets)} sheets: {sheets}")
        return sheets
    finally:
        if wb is not None:
            wb.close()


def parse_sheet(
    filepath: str | Path,
    sheet: str,
    *,
    header_row: int | None = None,
    data_start_row: int | None = None,
    data_end_row: int | None = None,
    start_col: str | None = None,
    end_col: str | None = None,
) -> tuple[list[str], list[dict[str, str | int | None]]]:
    """Read a sheet over a resolved row/column range, keyed by original headers.

    All bounds are optional and default per the module docstring (header on row
    1, data from ``header_row + 1`` to the last non-empty row in the span,
    columns from A to the first blank header). Every cell is coerced to a
    stripped string, empty -> ``None``. Merged ranges are resolved so
    continuation cells carry the anchor value; rows whose every value in the
    span is empty are skipped (logged with a count). Each kept data row dict
    carries a synthetic 1-based, contiguous ``row_number`` under
    ``ROW_NUMBER_KEY`` in addition to its header-keyed values.

    Args:
        filepath: Path to the Excel workbook.
        sheet: Name of the sheet to parse.
        header_row: 1-based header row. Defaults to 1.
        data_start_row: 1-based first data row. Defaults to ``header_row + 1``.
        data_end_row: 1-based last data row, inclusive. Defaults to the last row
            with content in the column span (best-effort).
        start_col: First column as an Excel letter. Defaults to ``"A"``.
        end_col: Last column as an Excel letter (inclusive). Defaults to the
            first blank header at or after ``start_col``.

    Returns:
        Tuple ``(column_names, rows)`` where ``column_names`` is the list of
        original header strings and ``rows`` is a list of dicts keyed by those
        headers plus ``ROW_NUMBER_KEY``.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the workbook cannot be opened, the sheet is missing, a
            column letter is invalid, the (resolved) row/column bounds are
            invalid, ``header_row`` is beyond the sheet, no headers are found,
            a header in an explicit column span is missing (blank), or two
            headers collide after cleaning/normalization.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    # Resolve the bounds that do not depend on the sheet content.
    header_row = _DEFAULT_HEADER_ROW if header_row is None else header_row
    if data_start_row is None:
        data_start_row = header_row + 1
    explicit_end = data_end_row is not None
    start_idx = _column_index(
        _DEFAULT_START_COL if start_col is None else start_col, "start_col"
    )
    end_idx = None if end_col is None else _column_index(end_col, "end_col")

    # Row numbers are 1-based everywhere in these configs; 0 or a negative
    # value is unambiguously a mistake. Without this check, header_row = 0
    # would compute header_idx = -1 and silently read the sheet's LAST row as
    # the header (Python negative indexing), ingesting the real header as data.
    if header_row < 1:
        raise ValueError(
            f"header_row ({header_row}) must be >= 1 (row numbers are 1-based)"
        )
    if data_start_row < 1:
        raise ValueError(
            f"data_start_row ({data_start_row}) must be >= 1 "
            "(row numbers are 1-based)"
        )
    if data_end_row is not None and data_end_row < 1:
        raise ValueError(
            f"data_end_row ({data_end_row}) must be >= 1 (row numbers are 1-based)"
        )
    if data_start_row <= header_row:
        raise ValueError(
            f"data_start_row ({data_start_row}) must be greater than "
            f"header_row ({header_row})"
        )
    if data_end_row is not None and data_end_row < data_start_row:
        raise ValueError(
            f"data_end_row ({data_end_row}) must be greater than or equal to "
            f"data_start_row ({data_start_row})"
        )
    if end_idx is not None and end_idx < start_idx:
        raise ValueError(
            f"end_col ({end_col}) must be at or after start_col "
            f"({_DEFAULT_START_COL if start_col is None else start_col})"
        )

    logger.info(
        f"Parsing sheet {sheet!r} from {filepath.name} "
        f"(header_row={header_row}, data_start_row={data_start_row}, "
        f"data_end_row={data_end_row if explicit_end else 'auto'}, "
        f"start_col={get_column_letter(start_idx + 1)}, "
        f"end_col={get_column_letter(end_idx + 1) if end_idx is not None else 'auto'})"
    )

    wb = None
    try:
        try:
            wb = load_workbook(filepath, read_only=True)
        except (OSError, ValueError, zipfile.BadZipFile, InvalidFileException) as e:
            # A corrupt/unreadable workbook can surface as BadZipFile or
            # InvalidFileException (neither a ValueError) and would otherwise
            # escape the orchestrator's per-sheet handler; wrap into ValueError
            # so it is recorded as a parse failure instead of aborting the run.
            logger.error(f"Failed to open workbook: {filepath} - {e}")
            raise ValueError(f"Cannot open workbook {filepath.name}: {e}") from e
        try:
            ws = wb[sheet]
        except KeyError as e:
            logger.error(f"Sheet {sheet!r} not found in {filepath.name}")
            raise ValueError(
                f"Sheet {sheet!r} not found in {filepath.name}"
            ) from e
        all_rows = list(ws.rows)
    finally:
        if wb is not None:
            wb.close()

    # Convert to plain values once, then resolve merged ranges: the read-only
    # reader yields the anchor (top-left) cell's value only, so every
    # continuation cell of a merged range parses as None — a vertically merged
    # category column would silently lose every continuation value. The merge
    # metadata is read from the sheet XML (read-only worksheets do not expose
    # merged_cells) and the anchor value is overlaid across each range.
    all_values = [[_cell_value(cell) for cell in row] for row in all_rows]
    _apply_merged_ranges(all_values, _merged_ranges(filepath, sheet))

    header_idx = header_row - 1
    if header_idx >= len(all_values):
        raise ValueError(
            f"header_row {header_row} is beyond the last row "
            f"({len(all_values)}) in sheet {sheet!r}"
        )

    raw_names = _extract_column_names(all_values[header_idx], start_idx, end_idx)
    _validate_headers(raw_names, sheet, header_row, start_idx)
    # _validate_headers rejected any None header, so this narrows losslessly.
    column_names: list[str] = [name for name in raw_names if name is not None]
    num_cols = len(column_names)
    logger.debug(f"Header at row {header_row}: {num_cols} columns")

    # Resolve an auto data_end_row to the last row with content in the span.
    if not explicit_end:
        data_end_row = _auto_data_end_row(
            all_values, data_start_row, start_idx, num_cols
        )
        logger.debug(f"Auto data_end_row resolved to {data_end_row}")

    rows: list[dict[str, str | int | None]] = []
    blank_rows_skipped = 0
    # Slice the resolved data range (0-based, end-inclusive). Rows beyond the
    # sheet's extent are silently absent from the slice.
    data_slice = all_values[data_start_row - 1 : data_end_row]
    for row_values in data_slice:
        values = [
            row_values[start_idx + j] if start_idx + j < len(row_values) else None
            for j in range(num_cols)
        ]
        # Skip an all-empty row: an interior blank row inside the data range
        # would otherwise become an all-None row dict that reaches
        # sheet_content (as a value-less row_text), the structured table, and
        # the source hash. The kept rows renumber contiguously so sort_order
        # stays 1..N.
        if all(value is None for value in values):
            blank_rows_skipped += 1
            continue
        row_dict: dict[str, str | int | None] = {ROW_NUMBER_KEY: len(rows) + 1}
        for name, value in zip(column_names, values):
            # Data cell for the j-th header sits at start_idx + j, so the data
            # is read from exactly the same column span as the headers.
            row_dict[name] = value
        rows.append(row_dict)

    if blank_rows_skipped:
        logger.info(
            f"Skipped {blank_rows_skipped} all-empty row(s) inside the data "
            f"range of sheet {sheet!r}"
        )
    logger.info(f"Parsed {sheet!r}: {len(rows)} data rows, {num_cols} columns")
    return column_names, rows


def _merged_ranges(
    filepath: Path, sheet: str
) -> list[tuple[int, int, int, int]]:
    """Read a sheet's merged-cell ranges straight from the workbook archive.

    openpyxl's read-only worksheets do not expose ``merged_cells``, and the
    read-only pass is kept for its memory profile — so the merge metadata is
    read from the sheet's XML part instead (an OOXML-standard location:
    ``xl/workbook.xml`` maps the sheet name to a relationship id,
    ``xl/_rels/workbook.xml.rels`` maps that id to the sheet part, whose
    ``<mergeCells>`` element lists the ranges).

    Best-effort: a workbook whose archive lacks these parts (a non-standard
    producer) simply reports no merges — the same behavior as before this
    resolution existed. A workbook broken enough to fail here would already
    have failed ``load_workbook`` in the caller.

    Args:
        filepath: Path to the ``.xlsx`` workbook.
        sheet: Name of the sheet whose merges to read.

    Returns:
        List of 1-based ``(min_col, min_row, max_col, max_row)`` bounds, one
        per merged range (openpyxl's ``range_boundaries`` order).
    """
    try:
        with zipfile.ZipFile(filepath) as archive:
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

            rel_id = None
            for sheet_el in workbook_root.iter(f"{_SPREADSHEET_NS}sheet"):
                if sheet_el.get("name") == sheet:
                    rel_id = sheet_el.get(_DOC_REL_ID)
                    break
            if rel_id is None:
                return []

            target = None
            for rel in rels_root.iter(f"{_PKG_REL_NS}Relationship"):
                if rel.get("Id") == rel_id:
                    target = rel.get("Target")
                    break
            if target is None:
                return []

            # Targets are workbook-relative unless rooted with '/'.
            part = (
                target.lstrip("/")
                if target.startswith("/")
                else posixpath.normpath(posixpath.join("xl", target))
            )
            sheet_root = ET.fromstring(archive.read(part))
    except (OSError, KeyError, zipfile.BadZipFile, ET.ParseError) as e:
        logger.warning(
            f"Could not read merged-cell metadata for sheet {sheet!r} in "
            f"{filepath.name} ({e}); continuation cells of merged ranges will "
            "parse as empty"
        )
        return []

    return [
        range_boundaries(merge_el.get("ref"))
        for merge_el in sheet_root.iter(f"{_SPREADSHEET_NS}mergeCell")
        if merge_el.get("ref")
    ]


def _apply_merged_ranges(
    all_values: list[list[str | None]],
    ranges: list[tuple[int, int, int, int]],
) -> None:
    """Overlay each merged range's anchor value across its continuation cells.

    Excel stores a merged range's value only in the anchor (top-left) cell, so
    the reader sees ``None`` for every other covered cell. Retrieval treats
    each row as a standalone record, so the anchor's value (e.g. a merged
    category label) is carried onto every spanned cell. Cells outside the
    materialized extent (a merge past the last populated row) are ignored.

    Args:
        all_values: The sheet's materialized cell values, mutated in place.
        ranges: 1-based ``(min_col, min_row, max_col, max_row)`` bounds.
    """
    for min_col, min_row, max_col, max_row in ranges:
        # Single-cell "ranges" carry nothing to spread.
        if min_row == max_row and min_col == max_col:
            continue
        anchor_row = min_row - 1
        anchor_col = min_col - 1
        if anchor_row >= len(all_values) or anchor_col >= len(all_values[anchor_row]):
            continue
        anchor_value = all_values[anchor_row][anchor_col]
        if anchor_value is None:
            continue
        for row_idx in range(min_row - 1, min(max_row, len(all_values))):
            row = all_values[row_idx]
            for col_idx in range(min_col - 1, max_col):
                if col_idx < len(row):
                    row[col_idx] = anchor_value


def _extract_column_names(
    header_values: list[str | None], start_idx: int, end_idx: int | None
) -> list[str | None]:
    """Extract header strings over the column span.

    For an explicit span (``end_idx`` not None) every column in
    ``[start_idx, end_idx]`` is returned, with ``None`` for a blank header (the
    caller rejects missing headers). For an auto span (``end_idx`` None) headers
    are read from ``start_idx`` until the first blank cell.

    Args:
        header_values: The header row's already-converted cell values.
        start_idx: 0-based first column of the span.
        end_idx: 0-based last column (inclusive), or None for auto.

    Returns:
        List of stripped header strings; contains ``None`` only for an explicit
        span with a blank header.
    """
    if end_idx is not None:
        return [
            header_values[i] if i < len(header_values) else None
            for i in range(start_idx, end_idx + 1)
        ]
    names: list[str | None] = []
    for i in range(start_idx, len(header_values)):
        val = header_values[i]
        if val is None:
            # An empty cell at/after start_col marks the end of the auto span.
            break
        names.append(val)
    return names


def _validate_headers(
    column_names: list[str | None], sheet: str, header_row: int, start_idx: int
) -> None:
    """Reject a sheet whose headers are missing or collide after cleaning.

    Args:
        column_names: Extracted headers over the span (may contain None).
        sheet: Sheet name (for messages).
        header_row: 1-based header row (for messages).
        start_idx: 0-based first column of the span (for column-letter messages).

    Raises:
        ValueError: If no headers are found, a header in the span is blank, or
            two headers are identical after cleaning or after normalization.
    """
    if not column_names:
        raise ValueError(
            f"No column headers found at row {header_row} starting at column "
            f"{get_column_letter(start_idx + 1)} in sheet {sheet!r}"
        )

    missing = [start_idx + j for j, n in enumerate(column_names) if n is None]
    if missing:
        cols = ", ".join(get_column_letter(i + 1) for i in missing)
        raise ValueError(
            f"Missing header(s) at column(s) {cols} in sheet {sheet!r} "
            "(blank cell in the configured column span); rejecting sheet"
        )

    # Duplicate after cleaning (stripped): identical raw headers collapse the
    # row dict, silently losing the earlier column's data.
    dup = _first_duplicate(column_names)
    if dup is not None:
        raise ValueError(
            f"Duplicate header {dup!r} in sheet {sheet!r} (identical after "
            "cleaning); rejecting sheet"
        )

    # Duplicate after normalization: two distinct headers map to the same col_*
    # structured-column name (e.g. 'Code' / 'code', 'CPT/HCPCS' / 'CPT HCPCS').
    # The missing-header check above guarantees no None survives here.
    normalized = [
        normalize_column_name(n, j)
        for j, n in enumerate(column_names)
        if n is not None
    ]
    ndup = _first_duplicate(normalized)
    if ndup is not None:
        raise ValueError(
            f"Headers in sheet {sheet!r} collide to the same column name "
            f"{ndup!r} after normalization; rejecting sheet"
        )


def _first_duplicate(items: Sequence[str | None]) -> str | None:
    """Return the first element that repeats in ``items``, or None if all unique."""
    seen: set[str | None] = set()
    for item in items:
        if item in seen:
            return item
        seen.add(item)
    return None


def _auto_data_end_row(
    all_values: list[list[str | None]],
    data_start_row: int,
    start_idx: int,
    num_cols: int,
) -> int:
    """Return the last 1-based row (>= data_start_row) with content in the span.

    Scans from ``data_start_row`` to the end of the sheet and returns the index
    of the last row that has any non-empty cell within the column span
    ``[start_idx, start_idx + num_cols)``. Returns ``data_start_row - 1`` (an
    empty range) when no such row exists.

    Best-effort: this includes any trailing footer / marker / total rows that
    carry content, so sheets with a footer below the data (e.g. an
    ``End of worksheet`` row) should set ``data_end_row`` explicitly.

    Args:
        all_values: The sheet's materialized, merge-resolved cell values.
        data_start_row: 1-based first data row.
        start_idx: 0-based first column of the span.
        num_cols: Number of columns in the span.

    Returns:
        The 1-based last data row, or ``data_start_row - 1`` if the span is empty.
    """
    last = data_start_row - 1
    end_excl = start_idx + num_cols
    for idx in range(data_start_row - 1, len(all_values)):
        row = all_values[idx]
        upper = min(end_excl, len(row))
        if any(row[c] is not None for c in range(start_idx, upper)):
            last = idx + 1  # 1-based
    return last


def _cell_value(cell: Cell) -> str | None:
    """Return a cell's stripped string value, or ``None`` if empty.

    Non-string cells (numbers, dates, booleans) are coerced via ``str()``, so
    they land in their Python string form (e.g. ``1.0``, a date's ISO-style
    repr, ``True``) rather than Excel's displayed format. This is intentional —
    the structured columns are all-text and ``row_text`` is a text rendering —
    but callers should not expect Excel's number/date formatting.

    Args:
        cell: An openpyxl cell.

    Returns:
        The stripped string value, or ``None`` for an empty/blank cell.
    """
    val = cell.value
    if val is None:
        return None
    text = str(val).strip()
    return text if text else None
