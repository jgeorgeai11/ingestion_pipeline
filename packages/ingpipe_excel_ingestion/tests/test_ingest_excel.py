"""Unit tests for the hybrid Excel ingestion orchestrator.

Config validation and the row_text builder are unit-tested directly. The
higher-level pipeline (collection_path identity, routing, overwrite/skip,
resilience) is exercised end-to-end against a real ephemeral schema where a real
DB is available, with in-memory workbooks; tests skip cleanly without a DB.
"""

import importlib
import os
import sys
from pathlib import Path

import pytest
from ingpipe_excel_ingestion import ingest_excel
from ingpipe_excel_ingestion._utils import compute_source_hash
from ingpipe_excel_ingestion.excel_parser import ROW_NUMBER_KEY
from ingpipe_excel_ingestion.ingest_excel import build_row_text, validate_config
from ingpipe_lib.logconfig import setup_logging
from ingpipe_lib.paths import resolve_log_dir
from ingpipe_lib.testing import assert_example_config_valid
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.engine import Engine

setup_logging(
    log_dir=resolve_log_dir("ingpipe_excel_ingestion/unit_tests"), log_name="test_ingest_excel"
)

# The four variables _utils.get_engine reads.
POSTGRES_VARS = ("POSTGRES_HOST", "POSTGRES_PORT", "POSTGRES_USER", "POSTGRES_PASSWORD")


def test_import_does_not_mutate_postgres_env(monkeypatch: pytest.MonkeyPatch) -> None:
    """Importing ingest_excel leaves the four POSTGRES_* variables alone.

    The module used to call ``load_dotenv()`` at module scope, so merely
    importing it populated the process environment from whatever ``.env`` the
    working directory happened to sit above. Credentials are now resolved
    inside ``main()`` from ``--env-file``, so a fresh import must be inert.

    This directory's ``conftest.py`` still performs its own ``load_dotenv()``
    for the ``ephemeral_schema`` fixture; that is deliberate and unaffected,
    which is why the variables are cleared here before the reload.
    """
    # Arrange: clear the four variables, then re-execute the module body.
    for var in POSTGRES_VARS:
        monkeypatch.delenv(var, raising=False)

    # Act
    importlib.reload(ingest_excel)

    # Assert: still unset, i.e. the import read no dotenv file.
    for var in POSTGRES_VARS:
        assert var not in os.environ


def test_main_without_env_file_flag_exits_usage_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    """A flag-less invocation is rejected by argparse (usage error, exit 2)."""
    # Arrange: argparse rejects before the config is ever opened, so the
    # config path only needs to exist as an argument.
    monkeypatch.setattr(
        sys, "argv", ["ingest_excel.py", "--config", str(tmp_path / "any.toml")]
    )

    # Act
    with pytest.raises(SystemExit) as exc:
        ingest_excel.main()

    # Assert: argparse's usage error names the missing flag.
    assert exc.value.code == 2
    assert "--env-file" in capsys.readouterr().err


def test_main_cli_overwrite_flag_overrides_toml_default(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, mocker
) -> None:
    """The --overwrite CLI flag overrides the TOML value (CLI-over-TOML).

    The config says overwrite = false; the flag flips both legs' write calls
    to overwrite=True. Everything DB-shaped is mocked so no engine is built.
    """
    # Arrange: minimal config with one embed-only sheet; source_dir absolute
    # so no instance root is required.
    src = tmp_path / "src"
    src.mkdir()
    config_path = tmp_path / "config.toml"
    config_path.write_text(
        f'source_dir = "{src.as_posix()}"\n'
        'db_name = "ingestion_test"\n'
        'db_schema = "test_schema"\n'
        "overwrite = false\n"
        '[files."wb.xlsx"]\n'
        'sheets = [\n  { sheet = "Alpha" },\n]\n',
        encoding="utf-8",
    )
    env_path = tmp_path / ".env.empty"
    env_path.write_text("", encoding="utf-8")
    monkeypatch.setattr(
        sys,
        "argv",
        ["ingest_excel.py", "--config", str(config_path), "--overwrite",
         "--env-file", str(env_path)],
    )
    mocker.patch("ingpipe_excel_ingestion.ingest_excel.get_engine")
    mocker.patch("ingpipe_excel_ingestion.ingest_excel.require_extensions")
    mocker.patch("ingpipe_excel_ingestion.ingest_excel.ensure_consolidated_tables")
    mocker.patch(
        "ingpipe_excel_ingestion.ingest_excel.parse_sheet",
        return_value=(["Code"], [{ROW_NUMBER_KEY: 1, "Code": "A1"}]),
    )
    mock_write = mocker.patch(
        "ingpipe_excel_ingestion.ingest_excel.write_consolidated", return_value=1
    )

    # Act
    ingest_excel.main()

    # Assert: the flag's True beat the config's false.
    assert mock_write.call_args.kwargs["overwrite"] is True


# ---------------------------------------------------------------------------
# validate_config
# ---------------------------------------------------------------------------


def _valid_config() -> dict:
    """A well-formed config with one both-legs sheet and one embed-only sheet."""
    return {
        "source_dir": "data/input/",
        "db_name": "test_db",
        "db_schema": "data",
        "files": {
            "wb.xlsx": {
                "sheets": [
                    {
                        "sheet": "A",
                        "header_row": 1,
                        "data_start_row": 2,
                        "data_end_row": 10,
                        "table": "sheet_a",
                    },
                    {
                        "sheet": "B",
                        "header_row": 1,
                        "data_start_row": 2,
                        "data_end_row": 5,
                    },
                ]
            }
        },
    }


def test_validate_config_accepts_valid() -> None:
    """A well-formed config passes."""
    validate_config(_valid_config())


def test_the_shipped_example_config_still_satisfies_validate_config() -> None:
    """The annotated config/example.toml this package ships still validates.

    Nothing ever executes the example, so drift from ``validate_config`` is
    silent and permanent: a required key added to the validator leaves a
    documented example that no longer works, and the cost lands on whoever
    copies it. This test is the only thing that runs it.
    """
    assert_example_config_valid("ingpipe_excel_ingestion", validate_config)


@pytest.mark.parametrize(
    "missing", ["source_dir", "db_name", "db_schema", "files"]
)
def test_validate_config_missing_top_level(missing: str) -> None:
    """A missing top-level field raises."""
    config = _valid_config()
    del config[missing]
    with pytest.raises(ValueError, match=f"Missing required config field.*{missing}"):
        validate_config(config)


def test_validate_config_missing_sheet_field() -> None:
    """The only required per-sheet field is `sheet`; omitting it raises."""
    config = _valid_config()
    del config["files"]["wb.xlsx"]["sheets"][0]["sheet"]
    with pytest.raises(ValueError, match="missing required"):
        validate_config(config)


def test_validate_config_row_bounds_all_optional() -> None:
    """A sheet entry with only `sheet` (no row/col bounds) is valid."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"] = [{"sheet": "A"}]
    validate_config(config)  # should not raise


@pytest.mark.parametrize("bad_value", [0, -1])
@pytest.mark.parametrize(
    "row_field", ["header_row", "data_start_row", "data_end_row"]
)
def test_validate_config_out_of_range_row_bound_raises(
    row_field: str, bad_value: int
) -> None:
    """A 0 or negative row bound is a clean config error (1-based rows).

    Without the range check a header_row = 0 would reach the parser as a
    negative index and silently read the wrong row.
    """
    config = _valid_config()
    # Replace the whole entry so the pairwise relationship checks (which run
    # only for explicitly provided pairs) cannot fire first.
    config["files"]["wb.xlsx"]["sheets"] = [{"sheet": "A", row_field: bad_value}]
    with pytest.raises(ValueError, match="1-based"):
        validate_config(config)


@pytest.mark.parametrize("col_field", ["start_col", "end_col"])
def test_validate_config_bad_column_letter_raises(col_field: str) -> None:
    """A non-Excel-letter column value aborts config validation."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0][col_field] = "B2"
    with pytest.raises(ValueError, match="Excel column letter"):
        validate_config(config)


def test_validate_config_accepts_valid_column_span() -> None:
    """A valid start_col/end_col pair passes."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["start_col"] = "B"
    config["files"]["wb.xlsx"]["sheets"][0]["end_col"] = "F"
    validate_config(config)  # should not raise


def test_validate_config_end_col_before_start_col_raises() -> None:
    """end_col earlier than start_col aborts config validation."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["start_col"] = "C"
    config["files"]["wb.xlsx"]["sheets"][0]["end_col"] = "A"
    with pytest.raises(ValueError, match="at or after start_col"):
        validate_config(config)


def test_validate_config_partial_bounds_valid() -> None:
    """A sheet with only some bounds (e.g. header_row alone) is valid."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"] = [{"sheet": "A", "header_row": 3}]
    validate_config(config)  # should not raise


def test_validate_config_table_is_optional() -> None:
    """The per-sheet table field is optional (embed-only sheet is valid)."""
    config = _valid_config()
    del config["files"]["wb.xlsx"]["sheets"][0]["table"]
    validate_config(config)  # should not raise


def test_validate_config_accepts_optional_title_and_collection_path() -> None:
    """Optional title and a valid authored collection_path pass."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["title"] = "A title"
    config["files"]["wb.xlsx"]["sheets"][0]["collection_path"] = "data.wb.a"
    validate_config(config)  # should not raise


def test_validate_config_invalid_collection_path_raises() -> None:
    """An authored collection_path that is not a valid ltree aborts validation."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["collection_path"] = "Bad Path"
    with pytest.raises(ValueError, match="Invalid collection_path"):
        validate_config(config)


def test_validate_config_accepts_a_valid_collection_path_prefix() -> None:
    """A valid ltree prefix passes config validation."""
    config = _valid_config()
    config["collection_path_prefix"] = "qpp_cm.2026_cost_measure_codes_lists"
    validate_config(config)  # should not raise


def test_validate_config_non_string_collection_path_prefix_raises() -> None:
    """A wrong-typed prefix is a clean config abort, not a per-sheet failure."""
    config = _valid_config()
    config["collection_path_prefix"] = 7
    with pytest.raises(ValueError, match="collection_path_prefix must be a string"):
        validate_config(config)


def test_validate_config_invalid_collection_path_prefix_raises() -> None:
    """A prefix that is not a valid ltree fails at load, not once per sheet."""
    config = _valid_config()
    config["collection_path_prefix"] = "Bad Prefix"
    with pytest.raises(ValueError, match="Invalid collection_path"):
        validate_config(config)


def test_validate_config_bad_row_bounds_start_not_after_header() -> None:
    """data_start_row <= header_row is a config error."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["data_start_row"] = 1
    with pytest.raises(ValueError, match="must be greater than header_row"):
        validate_config(config)


def test_validate_config_bad_row_bounds_end_before_start() -> None:
    """data_end_row < data_start_row is a config error."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["data_end_row"] = 1
    with pytest.raises(ValueError, match="must be >= data_start_row"):
        validate_config(config)


def test_validate_config_unsafe_table_identifier() -> None:
    """An unsafe table identifier aborts config validation."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["table"] = "Bad-Table"
    with pytest.raises(ValueError, match="Unsafe SQL identifier"):
        validate_config(config)


def test_validate_config_accepts_custom_consolidated_table_names() -> None:
    """Valid custom sheet_table/content_table names pass config validation."""
    config = _valid_config()
    config["sheet_table"] = "xls_sheet"
    config["content_table"] = "xls_sheet_content"
    validate_config(config)  # should not raise


@pytest.mark.parametrize("table_key", ["sheet_table", "content_table"])
def test_validate_config_unsafe_consolidated_table_name_raises(
    table_key: str,
) -> None:
    """An unsafe sheet_table/content_table name aborts config validation."""
    config = _valid_config()
    config[table_key] = "Bad-Name"
    with pytest.raises(ValueError, match=f"Unsafe SQL identifier for {table_key}"):
        validate_config(config)


def test_validate_config_equal_consolidated_table_names_raises() -> None:
    """sheet_table == content_table is rejected as a config error.

    An equal pair would make the second create-if-not-exists a no-op, hiding
    the missing content columns.
    """
    config = _valid_config()
    config["sheet_table"] = "both"
    config["content_table"] = "both"
    with pytest.raises(ValueError, match="must differ"):
        validate_config(config)


def test_validate_config_missing_sheets_key() -> None:
    """A file entry without a sheets list raises."""
    config = _valid_config()
    config["files"]["wb.xlsx"] = {}
    with pytest.raises(ValueError, match="missing required field 'sheets'"):
        validate_config(config)


def test_validate_config_non_int_row_bound_raises() -> None:
    """A quoted/non-int row bound aborts cleanly (ValueError), not TypeError."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["header_row"] = "1"
    with pytest.raises(ValueError, match="must be an integer"):
        validate_config(config)


def test_validate_config_sheets_not_a_list_raises() -> None:
    """A non-list sheets value aborts cleanly rather than failing mid-iteration."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"] = "oops"
    with pytest.raises(ValueError, match="must be a list"):
        validate_config(config)


@pytest.mark.parametrize(
    ("keys", "bad_value"),
    [
        (("files",), []),
        (("files", "wb.xlsx"), "oops"),
        (("files", "wb.xlsx", "sheets", 0), "oops"),
    ],
    ids=["files", "file_entry", "sheet_entry"],
)
def test_validate_config_non_table_entry_raises(
    keys: tuple[str | int, ...], bad_value: object
) -> None:
    """A non-table files / file entry / sheet entry aborts cleanly.

    Each of the three isinstance(..., dict) guards is reached by replacing the
    value at `keys` so the walk over the nested config cannot fail mid-iteration
    with an opaque TypeError instead.
    """
    config = _valid_config()
    target = config
    for key in keys[:-1]:
        target = target[key]
    target[keys[-1]] = bad_value
    with pytest.raises(ValueError, match="must be a table"):
        validate_config(config)


@pytest.mark.parametrize("bad", ["0.5", True, 1.5, -0.1])
def test_validate_config_bad_min_column_overlap_raises(bad: object) -> None:
    """A non-numeric, boolean, or out-of-range min_column_overlap aborts.

    `True` is an int subclass, so the explicit bool rejection is what keeps it
    from passing as 1; 1.5 and -0.1 cover the range check's two sides.
    """
    config = _valid_config()
    config["min_column_overlap"] = bad
    with pytest.raises(ValueError, match="min_column_overlap must be a number"):
        validate_config(config)


@pytest.mark.parametrize("good", [0.0, 0.5, 1.0])
def test_validate_config_accepts_min_column_overlap_bounds(good: float) -> None:
    """The inclusive [0, 1] boundaries and an interior value pass."""
    config = _valid_config()
    config["min_column_overlap"] = good
    validate_config(config)  # should not raise


def test_validate_config_non_bool_overwrite_raises() -> None:
    """A quoted overwrite aborts instead of silently taking the wrong branch.

    TOML `overwrite = "false"` parses to the truthy string "false", which would
    otherwise delete and re-write every present sheet — the opposite of intent.
    """
    config = _valid_config()
    config["overwrite"] = "false"
    with pytest.raises(ValueError, match="overwrite must be a boolean"):
        validate_config(config)


def test_validate_config_accepts_bool_overwrite() -> None:
    """A real boolean overwrite passes."""
    config = _valid_config()
    config["overwrite"] = True
    validate_config(config)  # should not raise


@pytest.mark.parametrize(
    "field", ["source_dir", "db_name", "db_schema", "sheet_table", "content_table"]
)
def test_validate_config_non_string_top_level_field_raises(field: str) -> None:
    """A non-string path / db / identifier field aborts with ValueError.

    Without the guard these reach Path(), get_engine(), or the identifier regex
    and raise TypeError, which escapes main()'s ValueError abort path.
    """
    config = _valid_config()
    config[field] = 5
    with pytest.raises(ValueError, match=f"{field} must be a string"):
        validate_config(config)


def test_validate_config_non_string_table_raises() -> None:
    """A non-string per-sheet table aborts with ValueError, not TypeError."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["table"] = 5
    with pytest.raises(ValueError, match="table must be a string"):
        validate_config(config)


def test_validate_config_accepts_comment_keys() -> None:
    """Valid top-level comment keys and a per-sheet table_comment pass."""
    config = _valid_config()
    config["schema_comment"] = "Schema text"
    config["sheet_table_comment"] = "Sheet catalog text"
    config["content_table_comment"] = "Row text"
    config["files"]["wb.xlsx"]["sheets"][0]["table_comment"] = "Structured text"
    validate_config(config)  # should not raise


@pytest.mark.parametrize(
    "comment_key",
    ["schema_comment", "sheet_table_comment", "content_table_comment"],
)
def test_validate_config_non_string_top_level_comment_raises(
    comment_key: str,
) -> None:
    """A non-string top-level comment value aborts config validation."""
    config = _valid_config()
    config[comment_key] = 42
    with pytest.raises(ValueError, match=f"{comment_key} must be a string"):
        validate_config(config)


def test_validate_config_table_comment_without_table_raises() -> None:
    """table_comment on an embed-only sheet (no table) is a config error."""
    config = _valid_config()
    # sheets[1] has no `table` key.
    config["files"]["wb.xlsx"]["sheets"][1]["table_comment"] = "orphaned"
    with pytest.raises(ValueError, match="table_comment requires a table"):
        validate_config(config)


def test_validate_config_non_string_table_comment_raises() -> None:
    """A non-string per-sheet table_comment aborts config validation."""
    config = _valid_config()
    config["files"]["wb.xlsx"]["sheets"][0]["table_comment"] = 7
    with pytest.raises(ValueError, match="table_comment must be a string"):
        validate_config(config)


# ---------------------------------------------------------------------------
# build_row_text (plain newline key-value)
# ---------------------------------------------------------------------------


def test_build_row_text_uses_original_headers_newline_kv() -> None:
    """row_text uses original headers, one `col: value` per line."""
    row = {"Item #": "1", "Name": "Widget", ROW_NUMBER_KEY: 1}
    assert build_row_text(["Item #", "Name"], row) == "Item #: 1\nName: Widget"


def test_build_row_text_none_becomes_empty() -> None:
    """None values render as empty strings."""
    row = {"A": "x", "B": None, ROW_NUMBER_KEY: 1}
    assert build_row_text(["A", "B"], row) == "A: x\nB: "


@pytest.mark.parametrize("falsy", [0, "", "0"])
def test_build_row_text_falsy_but_present_value_survives(falsy: str | int) -> None:
    """Only None renders as empty — a falsy-but-present value renders verbatim.

    Guards the `is None` check in build_row_text: a truthiness test would render
    an int 0 as "A: " and bake that loss into the source_binary_hash too.
    """
    row = {"A": falsy, ROW_NUMBER_KEY: 1}
    assert build_row_text(["A"], row) == f"A: {falsy}"


def test_build_row_text_excludes_row_number() -> None:
    """The synthetic row_number never appears in row_text."""
    row = {"A": "x", ROW_NUMBER_KEY: 99}
    assert "99" not in build_row_text(["A"], row)


# ---------------------------------------------------------------------------
# End-to-end pipeline (real ephemeral schema)
# ---------------------------------------------------------------------------


def _make_workbook(path: Path) -> None:
    """Write a workbook with a both-legs sheet and an embed-only sheet."""
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Alpha"
    ws_a.append(["Code", "Label"])
    ws_a.append(["A1", "alpha"])
    ws_a.append(["A2", "beta"])

    ws_b = wb.create_sheet("Bravo")
    ws_b.append(["Note"])
    ws_b.append(["just text"])

    wb.save(path)


def _run(config_path: Path) -> int:
    """Invoke ingest_excel.main() with --config and --env-file, returning the exit code."""
    # --env-file is required; an empty file beside the config satisfies the
    # flag while loading nothing, so the ambient (test) environment — already
    # pointed at ingestion_test by conftest — is untouched.
    env_path = config_path.parent / ".env.empty"
    env_path.write_text("", encoding="utf-8")
    argv = ["ingest_excel.py", "--config", str(config_path), "--env-file", str(env_path)]
    old = sys.argv
    sys.argv = argv
    try:
        ingest_excel.main()
        return 0
    except SystemExit as e:
        return int(e.code or 0)
    finally:
        sys.argv = old


def _write_config(
    path: Path, source_dir: Path, schema: str, sheets_toml: str, overwrite: bool
) -> None:
    """Write a minimal TOML config for the test workbook."""
    # as_posix(): a Windows tmp_path contains backslashes, which a TOML basic
    # string treats as escapes ("C:\Users" -> invalid \U hex escape); forward
    # slashes are valid Windows paths and need no escaping.
    path.write_text(
        f'source_dir = "{source_dir.as_posix()}"\n'
        'db_name = "ingestion_test"\n'
        f'db_schema = "{schema}"\n'
        f"overwrite = {str(overwrite).lower()}\n"
        '[files."wb.xlsx"]\n'
        f"sheets = [\n{sheets_toml}\n]\n",
        encoding="utf-8",
    )


def _content_rows(engine: Engine, schema: str, collection_path: str) -> list[dict]:
    """Fetch sheet_content rows for a collection_path ordered by sort_order."""
    with engine.connect() as conn:
        result = conn.execute(
            text(
                f'select * from "{schema}".sheet_content '
                "where collection_path = :cp order by sort_order"
            ),
            {"cp": collection_path},
        )
        return [dict(r._mapping) for r in result]


def _sheet_row(engine: Engine, schema: str, collection_path: str) -> dict | None:
    """Fetch the sheet metadata row for a collection_path, or None."""
    with engine.connect() as conn:
        result = conn.execute(
            text(
                f'select * from "{schema}".sheet where collection_path = :cp'
            ),
            {"cp": collection_path},
        ).fetchone()
        return dict(result._mapping) if result is not None else None


def test_pipeline_embed_only_and_both_legs(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """An embed-only sheet writes only sheet_content; a table sheet writes both."""
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")
    config = tmp_path / "cfg.toml"
    _write_config(
        config,
        src,
        schema,
        '    { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 3, table = "alpha" },\n'
        '    { sheet = "Bravo", header_row = 1, data_start_row = 2, '
        "data_end_row = 2 },",
        overwrite=False,
    )

    assert _run(config) == 0

    # Alpha: both legs (collection_path derives to wb.alpha).
    alpha_content = _content_rows(engine, schema, "wb.alpha")
    assert len(alpha_content) == 2
    # Pin the persisted row_text format end-to-end (newline key-value).
    assert alpha_content[0]["row_text"] == "Code: A1\nLabel: alpha"
    assert alpha_content[0]["word_count"] == 4
    with engine.connect() as conn:
        n = conn.execute(text(f'select count(*) from "{schema}".alpha')).scalar()
    assert n == 2

    # Bravo: embed-only — content present, no structured table named "bravo".
    assert len(_content_rows(engine, schema, "wb.bravo")) == 1
    with engine.connect() as conn:
        exists = conn.execute(
            text(
                "select 1 from information_schema.tables "
                "where table_schema = :s and table_name = 'bravo'"
            ),
            {"s": schema},
        ).fetchone()
    assert exists is None


def test_blank_sheet_title_rejected_by_postgres(
    ephemeral_schema: tuple[Engine, str],
) -> None:
    """The sheet table's trim(title) CHECK holds at rest.

    A config-authored whitespace-only title is rejected by PostgreSQL itself,
    mirroring ingpipe_file_ingestion's document.title check (schema symmetry).
    """
    from sqlalchemy.exc import IntegrityError

    engine, schema = ephemeral_schema
    ingest_excel.ensure_consolidated_tables(engine, schema)

    with pytest.raises(IntegrityError):
        with engine.begin() as conn:
            conn.execute(
                text(
                    f'insert into "{schema}".sheet '
                    "(collection_path, title, n_rows, source_binary_hash) "
                    "values ('wb.blank', '   ', 1, 1)"
                )
            )


def test_pipeline_stores_title_n_rows_and_hash(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """The sheet row records the default title, n_rows, and a per-sheet hash."""
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")
    config = tmp_path / "cfg.toml"
    _write_config(
        config,
        src,
        schema,
        '    { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 3, table = "alpha" },',
        overwrite=False,
    )

    assert _run(config) == 0
    row = _sheet_row(engine, schema, "wb.alpha")
    assert row is not None
    assert row["title"] == "wb — Alpha"  # default = "<workbook-stem> — <sheet>"
    assert row["n_rows"] == 2
    assert row["structured_table"] == "alpha"
    # Pin the per-sheet content fingerprint: the stored hash must equal the hash
    # of this sheet's newline-KV row_texts (a constant/0 implementation would not).
    expected_hash = compute_source_hash(
        ["Code: A1\nLabel: alpha", "Code: A2\nLabel: beta"]
    )
    assert int(row["source_binary_hash"]) == expected_hash
    assert expected_hash != 0


def test_pipeline_authored_collection_path_and_title(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """An authored collection_path + title override the derived defaults."""
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")
    config = tmp_path / "cfg.toml"
    _write_config(
        config,
        src,
        schema,
        '    { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 3, table = "alpha", title = "My Alpha", '
        'collection_path = "data.custom.alpha" },',
        overwrite=False,
    )

    assert _run(config) == 0
    assert _sheet_row(engine, schema, "wb.alpha") is None  # not the derived path
    row = _sheet_row(engine, schema, "data.custom.alpha")
    assert row is not None
    assert row["title"] == "My Alpha"
    assert len(_content_rows(engine, schema, "data.custom.alpha")) == 2


def test_pipeline_skip_if_present_then_overwrite(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """overwrite=false skips a present source on both legs; overwrite=true replaces."""
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")
    sheets = (
        '    { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 3, table = "alpha" },'
    )

    cfg1 = tmp_path / "cfg1.toml"
    _write_config(cfg1, src, schema, sheets, overwrite=False)
    assert _run(cfg1) == 0
    assert len(_content_rows(engine, schema, "wb.alpha")) == 2
    before = _sheet_row(engine, schema, "wb.alpha")

    # Re-run with overwrite=false -> skipped WITHOUT re-inserting the metadata row.
    assert _run(cfg1) == 0
    assert len(_content_rows(engine, schema, "wb.alpha")) == 2
    after = _sheet_row(engine, schema, "wb.alpha")
    # The sheet row must be untouched (ingested_at unchanged) — skip != re-insert.
    assert after == before

    # Now a shorter data range with overwrite=true -> replaced (1 row, both legs).
    cfg2 = tmp_path / "cfg2.toml"
    _write_config(
        cfg2,
        src,
        schema,
        '    { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 2, table = "alpha" },',
        overwrite=True,
    )
    assert _run(cfg2) == 0
    assert len(_content_rows(engine, schema, "wb.alpha")) == 1
    with engine.connect() as conn:
        n = conn.execute(text(f'select count(*) from "{schema}".alpha')).scalar()
    assert n == 1


def test_pipeline_skips_zero_row_sheet(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """A sheet whose range yields no rows is skipped on both legs (no orphan meta)."""
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    # Workbook with a header but no data in the configured range.
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    ws.append(["Code"])  # row 1 header only
    wb.save(src / "wb.xlsx")

    config = tmp_path / "cfg.toml"
    # data_start_row=2..data_end_row=2 is past the sheet extent -> 0 rows.
    config.write_text(
        f'source_dir = "{src.as_posix()}"\n'
        'db_name = "ingestion_test"\n'
        f'db_schema = "{schema}"\n'
        '[files."wb.xlsx"]\n'
        'sheets = [ { sheet = "Empty", header_row = 1, data_start_row = 2, '
        'data_end_row = 2, table = "empty" } ]\n'
    )

    assert _run(config) == 0
    # No sheet metadata row and no structured table were created.
    with engine.connect() as conn:
        meta = conn.execute(
            text(f'select count(*) from "{schema}".sheet')
        ).scalar()
        tbl = conn.execute(
            text(
                "select 1 from information_schema.tables "
                "where table_schema = :s and table_name = 'empty'"
            ),
            {"s": schema},
        ).fetchone()
    assert meta == 0
    assert tbl is None


def test_pipeline_resilient_mixed_batch(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """One sheet failing the structured guard still lets others succeed; exit 1."""
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()

    # Pre-seed structured table "alpha" with disjoint columns so the pipeline's
    # Alpha sheet clashes. The structured FK needs a parent sheet row first.
    from ingpipe_excel_ingestion.structured_table import write_rows

    ingest_excel.ensure_consolidated_tables(engine, schema)
    with engine.begin() as conn:
        conn.execute(
            text(
                f'insert into "{schema}".sheet '
                "(collection_path, title, n_rows, source_binary_hash) "
                "values ('other.x', 'X', 1, 0)"
            )
        )
    write_rows(
        engine, schema, "alpha", "sheet", "other.x",
        column_names=["Totally", "Different"],
        rows=[{ROW_NUMBER_KEY: 1, "Totally": "1", "Different": "2"}],
        overwrite=False, min_column_overlap=0.5,
    )

    _make_workbook(src / "wb.xlsx")
    config = tmp_path / "cfg.toml"
    _write_config(
        config,
        src,
        schema,
        '    { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 3, table = "alpha" },\n'
        '    { sheet = "Bravo", header_row = 1, data_start_row = 2, '
        "data_end_row = 2 },",
        overwrite=False,
    )

    # Alpha's structured write clashes -> failure recorded, non-zero exit.
    assert _run(config) == 1
    # Bravo (embed-only) still succeeded.
    assert len(_content_rows(engine, schema, "wb.bravo")) == 1
    # Alpha's embed leg succeeded before the structured clash (separate txns).
    assert len(_content_rows(engine, schema, "wb.alpha")) == 2


def test_pipeline_resilient_parse_failure(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """A sheet that fails to parse is recorded; other sheets still load; exit 1."""
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")
    config = tmp_path / "cfg.toml"
    _write_config(
        config,
        src,
        schema,
        '    { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 3, table = "alpha" },\n'
        '    { sheet = "DoesNotExist", header_row = 1, data_start_row = 2, '
        "data_end_row = 3 },",
        overwrite=False,
    )

    # The missing sheet fails parse_sheet -> recorded failure -> exit 1; Alpha
    # (a different sheet) still loads — the parse stage of the resilience contract.
    assert _run(config) == 1
    assert len(_content_rows(engine, schema, "wb.alpha")) == 2


def _table_exists(engine: Engine, schema: str, table: str) -> bool:
    """Return whether a table exists in the given schema."""
    with engine.connect() as conn:
        return (
            conn.execute(
                text(
                    "select 1 from information_schema.tables "
                    "where table_schema = :s and table_name = :t"
                ),
                {"s": schema, "t": table},
            ).fetchone()
            is not None
        )


def _object_comment(engine: Engine, schema: str, table: str | None) -> str | None:
    """Return the COMMENT ON description of a schema (table=None) or table."""
    with engine.connect() as conn:
        if table is None:
            return conn.execute(
                text(
                    "select obj_description(oid, 'pg_namespace') "
                    "from pg_namespace where nspname = :s"
                ),
                {"s": schema},
            ).scalar_one()
        return conn.execute(
            text(
                "select obj_description(c.oid, 'pg_class') "
                "from pg_class c "
                "inner join pg_namespace n on n.oid = c.relnamespace "
                "where n.nspname = :s and c.relname = :t"
            ),
            {"s": schema, "t": table},
        ).scalar_one()


def _comment_config_toml(
    source_dir: Path,
    schema: str,
    *,
    schema_comment: str,
    sheet_table_comment: str,
    content_table_comment: str,
    table_comment: str,
) -> str:
    """Build a TOML config carrying all four COMMENT ON override keys.

    The single sheet is both-legs (table = "alpha") so the per-sheet
    table_comment has a structured table to describe, and overwrite = false so
    that re-running with new text exercises the comment-refresh path rather than
    a re-insert.

    Args:
        source_dir: Directory holding wb.xlsx.
        schema: Target ephemeral schema.
        schema_comment: COMMENT ON SCHEMA text.
        sheet_table_comment: COMMENT ON TABLE text for the sheet table.
        content_table_comment: COMMENT ON TABLE text for the content table.
        table_comment: COMMENT ON TABLE text for the structured table.

    Returns:
        The TOML document text.
    """
    return (
        f'source_dir = "{source_dir.as_posix()}"\n'
        'db_name = "ingestion_test"\n'
        f'db_schema = "{schema}"\n'
        "overwrite = false\n"
        f'schema_comment = "{schema_comment}"\n'
        f'sheet_table_comment = "{sheet_table_comment}"\n'
        f'content_table_comment = "{content_table_comment}"\n'
        '[files."wb.xlsx"]\n'
        'sheets = [ { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 3, table = "alpha", '
        f'table_comment = "{table_comment}" }} ]\n'
    )


def test_pipeline_comment_keys_reach_the_catalog(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """Top-level comment keys and per-sheet table_comment land as descriptions.

    The schema/sheet/content overrides win over the template's baked generic
    comments; the structured table carries its per-sheet text (with a single
    quote surviving the round-trip).
    """
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")
    config = tmp_path / "cfg.toml"
    config.write_text(
        _comment_config_toml(
            src,
            schema,
            schema_comment="Test corpus schema",
            sheet_table_comment="Flavored sheet catalog",
            content_table_comment="Flavored row text",
            table_comment="Alpha's structured code list",
        )
    )

    assert _run(config) == 0

    assert _object_comment(engine, schema, None) == "Test corpus schema"
    assert _object_comment(engine, schema, "sheet") == "Flavored sheet catalog"
    assert _object_comment(engine, schema, "sheet_content") == "Flavored row text"
    assert (
        _object_comment(engine, schema, "alpha")
        == "Alpha's structured code list"
    )


def test_pipeline_comment_keys_refresh_on_rerun(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """A re-run with new comment text overwrites the existing descriptions.

    Both ensure_consolidated_tables and structured_table.ensure_table document
    that COMMENT ON is applied on EVERY call, not only at creation, which makes
    an idempotent re-run (overwrite = false, so both legs skip the rows) the
    supported way to refresh descriptions. A create-time-only implementation
    would leave the first run's text in place and fail here.
    """
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")

    first = tmp_path / "cfg1.toml"
    first.write_text(
        _comment_config_toml(
            src,
            schema,
            schema_comment="First schema text",
            sheet_table_comment="First sheet text",
            content_table_comment="First content text",
            table_comment="First structured text",
        )
    )
    assert _run(first) == 0
    assert _object_comment(engine, schema, "sheet") == "First sheet text"

    # Same sheet, same collection_path, overwrite = false -> the row writes are
    # skipped while the COMMENT ON statements still run with the new text.
    second = tmp_path / "cfg2.toml"
    second.write_text(
        _comment_config_toml(
            src,
            schema,
            schema_comment="Second schema text",
            sheet_table_comment="Second sheet text",
            content_table_comment="Second content text",
            table_comment="Second structured text",
        )
    )
    assert _run(second) == 0

    assert _object_comment(engine, schema, None) == "Second schema text"
    assert _object_comment(engine, schema, "sheet") == "Second sheet text"
    assert _object_comment(engine, schema, "sheet_content") == "Second content text"
    assert _object_comment(engine, schema, "alpha") == "Second structured text"
    # The refresh must not have re-written the skipped sheet's rows.
    assert len(_content_rows(engine, schema, "wb.alpha")) == 2


def test_pipeline_template_bakes_generic_comments_without_overrides(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """Without comment keys, the template's generic table comments still apply."""
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")
    config = tmp_path / "cfg.toml"
    _write_config(
        config,
        src,
        schema,
        '    { sheet = "Bravo", header_row = 1, data_start_row = 2, '
        "data_end_row = 2 },",
        overwrite=False,
    )

    assert _run(config) == 0

    sheet_comment = _object_comment(engine, schema, "sheet")
    content_comment = _object_comment(engine, schema, "sheet_content")
    assert sheet_comment is not None
    assert "Catalog of ingested Excel worksheets" in sheet_comment
    assert content_comment is not None
    assert "row-level text" in content_comment
    # No schema_comment key -> the schema description is left alone (None here).
    assert _object_comment(engine, schema, None) is None


def test_pipeline_custom_consolidated_table_names(
    ephemeral_schema: tuple[Engine, str], tmp_path: Path
) -> None:
    """Custom sheet_table/content_table names flow into the DDL, inserts, and FK.

    Verifies the configured names reach the rendered DDL (the custom tables are
    created) and the insert/select SQL (they hold rows), that the default
    sheet/sheet_content tables are NOT created, and that the structured leg's FK
    references the CUSTOM sheet table (it is created — an absent FK target would
    fail at create time). An overwrite re-run exercises the
    _content_source_present select and the content cascade against the custom
    name (the structured row is re-written by write_rows's own delete-insert).
    """
    engine, schema = ephemeral_schema
    src = tmp_path / "src"
    src.mkdir()
    _make_workbook(src / "wb.xlsx")
    config = tmp_path / "cfg.toml"
    # Inline config (not _write_config) so the custom top-level table names are set.
    config.write_text(
        f'source_dir = "{src.as_posix()}"\n'
        'db_name = "ingestion_test"\n'
        f'db_schema = "{schema}"\n'
        'sheet_table = "xls_sheet"\n'
        'content_table = "xls_sheet_content"\n'
        "overwrite = false\n"
        '[files."wb.xlsx"]\n'
        'sheets = [ { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 3, table = "alpha" } ]\n'
    )

    assert _run(config) == 0

    # The custom consolidated tables were created and hold rows; the structured
    # table's FK to the custom sheet table held (alpha row inserted).
    assert _table_exists(engine, schema, "xls_sheet")
    assert _table_exists(engine, schema, "xls_sheet_content")
    with engine.connect() as conn:
        meta = conn.execute(
            text(
                f'select count(*) from "{schema}".xls_sheet '
                "where collection_path = 'wb.alpha'"
            )
        ).scalar()
        content = conn.execute(
            text(
                f'select count(*) from "{schema}".xls_sheet_content '
                "where collection_path = 'wb.alpha'"
            )
        ).scalar()
        structured = conn.execute(
            text(f'select count(*) from "{schema}".alpha')
        ).scalar()
    assert meta == 1
    assert content == 2
    assert structured == 2

    # The default-named tables must NOT have been created in this schema.
    assert not _table_exists(engine, schema, "sheet")
    assert not _table_exists(engine, schema, "sheet_content")

    # Re-run with overwrite=true and a shorter range: exercises the
    # _content_source_present select against the custom sheet table, the content
    # cascade (custom sheet delete cascades to custom content), and the
    # structured leg's own delete-reinsert (write_rows manages its rows).
    config2 = tmp_path / "cfg2.toml"
    config2.write_text(
        f'source_dir = "{src.as_posix()}"\n'
        'db_name = "ingestion_test"\n'
        f'db_schema = "{schema}"\n'
        'sheet_table = "xls_sheet"\n'
        'content_table = "xls_sheet_content"\n'
        "overwrite = true\n"
        '[files."wb.xlsx"]\n'
        'sheets = [ { sheet = "Alpha", header_row = 1, data_start_row = 2, '
        'data_end_row = 2, table = "alpha" } ]\n'
    )
    assert _run(config2) == 0
    with engine.connect() as conn:
        content_after = conn.execute(
            text(
                f'select count(*) from "{schema}".xls_sheet_content '
                "where collection_path = 'wb.alpha'"
            )
        ).scalar()
        structured_after = conn.execute(
            text(f'select count(*) from "{schema}".alpha')
        ).scalar()
    assert content_after == 1
    assert structured_after == 1
