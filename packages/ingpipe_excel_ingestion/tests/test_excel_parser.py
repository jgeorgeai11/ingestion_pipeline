"""Unit tests for the explicit-row-bound Excel parser."""

from pathlib import Path

import pytest
from ingpipe_excel_ingestion.excel_parser import ROW_NUMBER_KEY, list_sheets, parse_sheet
from ingpipe_lib.logconfig import setup_logging
from ingpipe_lib.paths import resolve_log_dir
from openpyxl import Workbook

setup_logging(
    log_dir=resolve_log_dir("ingpipe_excel_ingestion/unit_tests"), log_name="test_excel_parser"
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def sample_workbook(tmp_path: Path) -> Path:
    """A workbook with a header-on-row-1 sheet and an offset-header sheet."""
    filepath = tmp_path / "sample.xlsx"
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Products"
    ws1.append(["Product ID", "Product Name", "Unit Price"])  # row 1
    ws1.append(["P001", "Widget A", "9.99"])                  # row 2
    ws1.append(["P002", "Widget B", "14.50"])                 # row 3
    ws1.append(["P003", "Widget C", "22.00"])                 # row 4
    ws1.append(["P004", "Widget D", "30.00"])                 # row 5 (outside bound)

    ws2 = wb.create_sheet("Orders")
    ws2.append(["Order Report - Q1 2026"])     # row 1 (title)
    ws2.append(["Order #", "Customer", "Amount"])  # row 2 (header)
    ws2.append(["1001", "Acme Corp", "500.00"])    # row 3
    ws2.append(["1002", "Globex", "750.00"])       # row 4

    wb.save(filepath)
    return filepath


@pytest.fixture()
def gappy_header_workbook(tmp_path: Path) -> Path:
    """A workbook whose header row has a gap then a stray later header."""
    filepath = tmp_path / "gappy.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Gappy"
    # Header set must terminate at the first empty cell, ignoring "Stray".
    ws.append(["Col A", "Col B", None, "Stray"])  # row 1
    ws.append(["a1", "b1", "x", "y"])             # row 2
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# list_sheets
# ---------------------------------------------------------------------------


def test_list_sheets_returns_all_names(sample_workbook: Path) -> None:
    """list_sheets returns every sheet name in workbook order."""
    assert list_sheets(sample_workbook) == ["Products", "Orders"]


def test_list_sheets_file_not_found(tmp_path: Path) -> None:
    """list_sheets raises FileNotFoundError for a missing file."""
    with pytest.raises(FileNotFoundError, match="Excel file not found"):
        list_sheets(tmp_path / "nope.xlsx")


def test_list_sheets_corrupt_file_raises_value_error(tmp_path: Path) -> None:
    """A non-xlsx file with an .xlsx name surfaces as a wrapped ValueError."""
    bad = tmp_path / "bad.xlsx"
    bad.write_text("this is not a real xlsx zip")
    with pytest.raises(ValueError, match="Cannot open workbook"):
        list_sheets(bad)


# ---------------------------------------------------------------------------
# parse_sheet: happy paths
# ---------------------------------------------------------------------------


def test_parse_sheet_reads_headers_and_bounded_rows(sample_workbook: Path) -> None:
    """Headers locate columns; data is read over the inclusive bound only."""
    columns, rows = parse_sheet(
        sample_workbook,
        "Products",
        header_row=1,
        data_start_row=2,
        data_end_row=4,
    )

    assert columns == ["Product ID", "Product Name", "Unit Price"]
    assert len(rows) == 3  # rows 2-4 inclusive; row 5 excluded
    assert rows[0]["Product ID"] == "P001"
    assert rows[2]["Product ID"] == "P003"
    # Row 5 (P004) is outside the bound.
    assert all(r["Product ID"] != "P004" for r in rows)


def test_parse_sheet_offset_header(sample_workbook: Path) -> None:
    """A header not on row 1 is located by header_row."""
    columns, rows = parse_sheet(
        sample_workbook,
        "Orders",
        header_row=2,
        data_start_row=3,
        data_end_row=4,
    )
    assert columns == ["Order #", "Customer", "Amount"]
    assert len(rows) == 2
    assert rows[0]["Customer"] == "Acme Corp"


def test_parse_sheet_synthetic_row_number(sample_workbook: Path) -> None:
    """Each data row carries a 1-based synthetic row_number."""
    _, rows = parse_sheet(
        sample_workbook,
        "Products",
        header_row=1,
        data_start_row=2,
        data_end_row=4,
    )
    assert [r[ROW_NUMBER_KEY] for r in rows] == [1, 2, 3]
    # row_number is keyed separately from the header columns: pin against the
    # actual row keys, not string literals (which can never fail).
    assert set(rows[0]) == {
        "Product ID",
        "Product Name",
        "Unit Price",
        ROW_NUMBER_KEY,
    }


def test_parse_sheet_blank_cells_become_none(tmp_path: Path) -> None:
    """Empty cells coerce to None; values coerce to stripped strings."""
    filepath = tmp_path / "blanks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B", "C"])
    ws.append(["x", None, 5])  # numeric coerces to str; blank -> None
    wb.save(filepath)

    _, rows = parse_sheet(
        filepath, "S", header_row=1, data_start_row=2, data_end_row=2
    )
    assert rows[0]["A"] == "x"
    assert rows[0]["B"] is None
    assert rows[0]["C"] == "5"
    assert isinstance(rows[0]["C"], str)


def test_parse_sheet_header_gap_terminates_header_set(
    gappy_header_workbook: Path,
) -> None:
    """The header set terminates at the first empty header cell."""
    columns, rows = parse_sheet(
        gappy_header_workbook,
        "Gappy",
        header_row=1,
        data_start_row=2,
        data_end_row=2,
    )
    assert columns == ["Col A", "Col B"]  # "Stray" past the gap is ignored
    assert set(rows[0]) == {"Col A", "Col B", ROW_NUMBER_KEY}


# ---------------------------------------------------------------------------
# parse_sheet: validation / error paths
# ---------------------------------------------------------------------------


def test_parse_sheet_data_start_not_after_header_raises(
    sample_workbook: Path,
) -> None:
    """data_start_row must be greater than header_row."""
    with pytest.raises(ValueError, match="must be greater than"):
        parse_sheet(
            sample_workbook,
            "Products",
            header_row=2,
            data_start_row=1,
            data_end_row=4,
        )


def test_parse_sheet_data_end_before_start_raises(sample_workbook: Path) -> None:
    """data_end_row must be >= data_start_row."""
    with pytest.raises(ValueError, match="greater than or equal to"):
        parse_sheet(
            sample_workbook,
            "Products",
            header_row=1,
            data_start_row=4,
            data_end_row=2,
        )


def test_parse_sheet_vertically_merged_cells_carry_value(tmp_path: Path) -> None:
    """A vertically merged column yields the anchor value on every spanned row.

    Reproduces the row-fidelity defect: the read-only reader stores a merged
    range's value only in the anchor cell, so continuation rows parsed as
    None and every merged category value was silently lost.
    """
    filepath = tmp_path / "merged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws.append(["Group", "Val"])
    ws.append(["grp", "1"])
    ws.append([None, "2"])
    ws.append([None, "3"])
    ws.merge_cells("A2:A4")
    wb.save(filepath)

    columns, rows = parse_sheet(filepath, "Merged")

    assert columns == ["Group", "Val"]
    assert [r["Group"] for r in rows] == ["grp", "grp", "grp"]
    assert [r["Val"] for r in rows] == ["1", "2", "3"]


def test_parse_sheet_horizontally_merged_header_fills_span(tmp_path: Path) -> None:
    """A merged range never overwrites cells outside the materialized extent."""
    filepath = tmp_path / "merged_past_end.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws.append(["A", "B"])
    ws.append(["x", "y"])
    # Merge extending past the populated extent must not crash the overlay.
    ws.merge_cells("A2:A10")
    wb.save(filepath)

    columns, rows = parse_sheet(filepath, "M")

    assert columns == ["A", "B"]
    assert rows[0]["A"] == "x"


def test_parse_sheet_blank_interior_rows_skipped(tmp_path: Path) -> None:
    """Blank rows inside the data range are skipped, not ingested as all-None.

    Reproduces the row-fidelity defect: the auto data_end_row includes any
    trailing content, so intervening blank rows became all-None row dicts.
    The kept rows renumber contiguously (sort_order stays 1..N).
    """
    filepath = tmp_path / "gaps.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Gaps"
    ws.append(["Code", "Label"])
    ws.append(["A1", "alpha"])
    ws.append([None, None])       # interior blank row
    ws.append(["A2", "beta"])
    ws.append([None, None])       # another blank
    ws.append(["A3", "gamma"])
    wb.save(filepath)

    _columns, rows = parse_sheet(filepath, "Gaps")

    assert [r["Code"] for r in rows] == ["A1", "A2", "A3"]
    # Contiguous 1-based numbering over the KEPT rows only.
    assert [r[ROW_NUMBER_KEY] for r in rows] == [1, 2, 3]


@pytest.mark.parametrize("bad_value", [0, -1])
@pytest.mark.parametrize(
    "row_field", ["header_row", "data_start_row", "data_end_row"]
)
def test_parse_sheet_out_of_range_row_bound_raises(
    sample_workbook: Path, row_field: str, bad_value: int
) -> None:
    """A 0 or negative row bound is rejected, never index-wrapped.

    Reproduces the silent-data-loss defect: header_row = 0 became index -1 and
    silently used the sheet's LAST row as the header. All three 1-based row
    settings must reject values below 1 with a clear error.
    """
    kwargs: dict[str, int] = {row_field: bad_value}
    with pytest.raises(ValueError, match="1-based"):
        parse_sheet(sample_workbook, "Products", **kwargs)


def test_parse_sheet_invalid_sheet_name_raises(sample_workbook: Path) -> None:
    """A missing sheet raises ValueError."""
    with pytest.raises(ValueError, match="not found"):
        parse_sheet(
            sample_workbook,
            "NoSuchSheet",
            header_row=1,
            data_start_row=2,
            data_end_row=4,
        )


def test_parse_sheet_file_not_found_raises(tmp_path: Path) -> None:
    """A missing file raises FileNotFoundError."""
    with pytest.raises(FileNotFoundError, match="Excel file not found"):
        parse_sheet(
            tmp_path / "missing.xlsx",
            "S",
            header_row=1,
            data_start_row=2,
            data_end_row=4,
        )


def test_parse_sheet_keyword_only_row_args(sample_workbook: Path) -> None:
    """The row-bound arguments are keyword-only."""
    with pytest.raises(TypeError):
        parse_sheet(sample_workbook, "Products", 1, 2, 4)  # type: ignore[misc]


def test_parse_sheet_corrupt_file_raises_value_error(tmp_path: Path) -> None:
    """A corrupt/non-xlsx workbook surfaces as a wrapped ValueError (parse failure)."""
    bad = tmp_path / "bad.xlsx"
    bad.write_text("this is not a real xlsx zip")
    with pytest.raises(ValueError, match="Cannot open workbook"):
        parse_sheet(bad, "S", header_row=1, data_start_row=2, data_end_row=2)


def test_parse_sheet_no_headers_raises(tmp_path: Path) -> None:
    """An all-blank header row yields no columns and raises (documented Raises)."""
    filepath = tmp_path / "noheaders.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    # Whitespace cells are persisted but strip to None, so the header set is empty.
    ws.append([" ", " ", " "])  # row 1: blank header
    ws.append(["a", "b", "c"])  # row 2: data
    wb.save(filepath)

    with pytest.raises(ValueError, match="No column headers found"):
        parse_sheet(filepath, "S", header_row=1, data_start_row=2, data_end_row=2)


def test_parse_sheet_header_row_beyond_extent_raises(tmp_path: Path) -> None:
    """A header_row past the sheet's last populated row raises (documented Raises)."""
    filepath = tmp_path / "small.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])  # only row 1 is populated
    wb.save(filepath)

    with pytest.raises(ValueError, match="beyond the last row"):
        parse_sheet(filepath, "S", header_row=5, data_start_row=6, data_end_row=6)


def test_parse_sheet_ragged_row_pads_with_none(tmp_path: Path) -> None:
    """A data row shorter than the header pads the missing trailing columns to None."""
    filepath = tmp_path / "ragged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B", "C"])  # three headers
    ws.append(["1"])            # one cell -> B, C padded with None
    wb.save(filepath)

    _, rows = parse_sheet(
        filepath, "S", header_row=1, data_start_row=2, data_end_row=2
    )
    assert rows[0]["A"] == "1"
    assert rows[0]["B"] is None
    assert rows[0]["C"] is None


# ---------------------------------------------------------------------------
# parse_sheet: defaults + auto data_end_row
# ---------------------------------------------------------------------------


def test_parse_sheet_defaults(tmp_path: Path) -> None:
    """With no bounds, header is row 1, data from row 2, span auto from A."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code", "Label"])  # row 1 header
    ws.append(["A1", "alpha"])    # row 2
    ws.append(["A2", "beta"])     # row 3
    wb.save(filepath)

    cols, rows = parse_sheet(filepath, "S")
    assert cols == ["Code", "Label"]
    assert len(rows) == 2
    assert rows[0]["Code"] == "A1"
    assert rows[1]["Label"] == "beta"


def test_parse_sheet_auto_data_end_row(tmp_path: Path) -> None:
    """Auto data_end_row resolves to the last row with content in the span."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code"])
    for v in ["a", "b", "c", "d"]:
        ws.append([v])
    wb.save(filepath)

    _, rows = parse_sheet(filepath, "S")  # header row 1, data 2.., auto end
    assert [r["Code"] for r in rows] == ["a", "b", "c", "d"]


def test_parse_sheet_auto_data_end_row_includes_footer(tmp_path: Path) -> None:
    """Auto end is best-effort: a trailing footer row with content is included."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code"])
    ws.append(["a"])
    ws.append(["b"])
    ws.append(["End of worksheet"])  # footer with content, no blank gap
    wb.save(filepath)

    # The documented caveat: auto includes the footer (set data_end_row to exclude).
    _, rows = parse_sheet(filepath, "S")
    assert [r["Code"] for r in rows] == ["a", "b", "End of worksheet"]


# ---------------------------------------------------------------------------
# parse_sheet: column bounds
# ---------------------------------------------------------------------------


def test_parse_sheet_start_col_offsets_span(tmp_path: Path) -> None:
    """start_col reads headers and data from that column, not A."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["row label", "Code", "Label"])  # A is a non-header label
    ws.append(["r1", "A1", "alpha"])
    wb.save(filepath)

    cols, rows = parse_sheet(filepath, "S", data_end_row=2, start_col="B")
    assert cols == ["Code", "Label"]
    assert rows[0]["Code"] == "A1"
    assert rows[0]["Label"] == "alpha"
    assert "row label" not in rows[0]


def test_parse_sheet_end_col_limits_span(tmp_path: Path) -> None:
    """end_col bounds the span; columns beyond it are ignored."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code", "Label", "Ignored"])
    ws.append(["A1", "alpha", "junk"])
    wb.save(filepath)

    cols, rows = parse_sheet(filepath, "S", data_end_row=2, end_col="B")
    assert cols == ["Code", "Label"]
    assert "junk" not in rows[0].values()


def test_parse_sheet_bad_column_letter_raises(tmp_path: Path) -> None:
    """A non-letter column value is rejected."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code"])
    ws.append(["x"])
    wb.save(filepath)

    with pytest.raises(ValueError, match="Excel column letter"):
        parse_sheet(filepath, "S", data_end_row=2, start_col="1")


# ---------------------------------------------------------------------------
# parse_sheet: reject missing / duplicate headers
# ---------------------------------------------------------------------------


def test_parse_sheet_explicit_span_missing_header_raises(tmp_path: Path) -> None:
    """A blank header within an explicit column span is rejected."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code", None, "Label"])  # B header blank
    ws.append(["A1", "x", "alpha"])
    wb.save(filepath)

    with pytest.raises(ValueError, match="Missing header"):
        parse_sheet(filepath, "S", data_end_row=2, end_col="C")


def test_parse_sheet_rejects_duplicate_header(tmp_path: Path) -> None:
    """Two identical headers are rejected (would collapse, losing data)."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code", "Code"])
    ws.append(["A1", "A2"])
    wb.save(filepath)

    with pytest.raises(ValueError, match="identical after cleaning"):
        parse_sheet(filepath, "S", data_end_row=2)


def test_parse_sheet_rejects_normalize_collision(tmp_path: Path) -> None:
    """Two headers that normalize to the same col_* name are rejected."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code", "code"])  # distinct raw, both -> col_code
    ws.append(["A1", "A2"])
    wb.save(filepath)

    with pytest.raises(ValueError, match="after normalization"):
        parse_sheet(filepath, "S", data_end_row=2)


def test_parse_sheet_leading_blank_header_no_longer_misaligns(
    tmp_path: Path,
) -> None:
    """A blank header at the start column is a clean error, not a silent shift."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append([None, "Code", "Desc"])  # column A header blank
    ws.append(["aval", "bval", "cval"])
    wb.save(filepath)

    # Previously this silently shifted Code<-colA; now it's a clean error.
    with pytest.raises(ValueError, match="No column headers found"):
        parse_sheet(filepath, "S", data_end_row=2)

    # The fix: declare the real span with start_col -> correct alignment.
    cols, rows = parse_sheet(filepath, "S", data_end_row=2, start_col="B")
    assert cols == ["Code", "Desc"]
    assert rows[0]["Code"] == "bval"


def test_parse_sheet_end_col_before_start_col_raises(tmp_path: Path) -> None:
    """end_col earlier than start_col is rejected."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code", "Label"])
    ws.append(["a", "b"])
    wb.save(filepath)

    with pytest.raises(ValueError, match="at or after start_col"):
        parse_sheet(filepath, "S", data_end_row=2, start_col="C", end_col="A")


def test_parse_sheet_auto_end_no_data_rows(tmp_path: Path) -> None:
    """A header with no data below yields zero rows under auto data_end_row."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code"])  # header only, no data
    wb.save(filepath)

    cols, rows = parse_sheet(filepath, "S")  # auto end over an empty data range
    assert cols == ["Code"]
    assert rows == []


def test_parse_sheet_data_start_defaults_to_header_row_plus_one(
    tmp_path: Path,
) -> None:
    """With header_row set but data_start_row omitted, data starts at header+1."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["junk"])   # row 1
    ws.append(["junk2"])  # row 2
    ws.append(["Code"])   # row 3 header
    ws.append(["a"])      # row 4 data
    ws.append(["b"])      # row 5 data
    wb.save(filepath)

    cols, rows = parse_sheet(filepath, "S", header_row=3)  # data_start -> 4
    assert cols == ["Code"]
    assert [r["Code"] for r in rows] == ["a", "b"]


def test_parse_sheet_explicit_span_wider_than_data_rows(tmp_path: Path) -> None:
    """An explicit span wider than the data rows pads trailing columns to None."""
    filepath = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Code", "Label", "Extra"])  # 3 headers
    ws.append(["a", "b"])                   # data row has only 2 cells
    wb.save(filepath)

    _, rows = parse_sheet(filepath, "S", data_end_row=2, end_col="C")
    assert rows[0]["Code"] == "a"
    assert rows[0]["Label"] == "b"
    assert rows[0]["Extra"] is None
